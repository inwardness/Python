"""
ЧИСТОЕ КОПИРОВАНИЕ .wop ФАЙЛОВ В EXCEL
1. Открываем первый файл
2. Копируем ВСЁ содержимое как есть в первую строку Excel
3. Добавляем одну пустую строку
4. Открываем следующий файл
5. Копируем его содержимое после пустой строки
6. Повторяем для всех файлов
"""

import os
import glob
from openpyxl import Workbook

def read_file_all_content(file_path):
    """
    Читаем ВСЁ содержимое файла как есть
    """
    encodings = ['utf-8', 'windows-1251', 'cp1251', 'latin-1', 'cp866', 'ascii']
    
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as file:
                return file.read(), encoding
        except (UnicodeDecodeError, LookupError):
            continue
    
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            return file.read(), 'utf-8 (ignore errors)'
    except Exception as e:
        return f"ERROR READING FILE: {str(e)}", 'error'

def copy_wop_to_excel_simple():
    """
    Простое копирование: файл за файлом без заголовков
    """
    print("=" * 60)
    print("КОПИРОВАНИЕ .wop ФАЙЛОВ В EXCEL (ПРОСТОЙ РЕЖИМ)")
    print("=" * 60)
    
    # Ищем все .wop файлы
    wop_files = []
    
    # В текущей папке
    for f in glob.glob("*.wop"):
        wop_files.append(f)
    
    # В подпапках
    for root, dirs, files in os.walk("."):
        for file in files:
            if file.lower().endswith('.wop'):
                wop_files.append(os.path.join(root, file))
    
    if not wop_files:
        print("\n❌ Файлы .wop не найдены!")
        print("Положите .wop файлы в папку с программой или в подпапки")
        return None
    
    # Сортируем по имени
    wop_files.sort()
    
    print(f"\nНайдено файлов: {len(wop_files)}")
    for i, f in enumerate(wop_files, 1):
        print(f"{i:3}. {os.path.basename(f)}")
    
    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "WOP Files"
    
    current_row = 1
    
    print("\n" + "=" * 60)
    print("КОПИРУЮ СОДЕРЖИМОЕ...")
    print("=" * 60)
    
    # Обрабатываем каждый файл
    for file_idx, file_path in enumerate(wop_files, 1):
        file_name = os.path.basename(file_path)
        
        print(f"Файл {file_idx}/{len(wop_files)}: {file_name}")
        
        # Читаем содержимое
        content, encoding = read_file_all_content(file_path)
        
        # Разделяем на строки
        lines = content.split('\n')
        
        # Копируем строки в Excel
        for line in lines:
            ws.cell(row=current_row, column=1, value=line)
            current_row += 1
        
        # Добавляем одну пустую строку между файлами (кроме последнего)
        if file_idx < len(wop_files):
            current_row += 1
        
        print(f"  Строк: {len(lines)}, кодировка: {encoding}")
    
    # Настраиваем ширину
    ws.column_dimensions['A'].width = 150
    
    # Сохраняем
    output_file = "wop_files_combined.xlsx"
    wb.save(output_file)
    
    print("\n" + "=" * 60)
    print("✅ ГОТОВО!")
    print("=" * 60)
    print(f"Создан файл: {output_file}")
    print(f"Всего строк в Excel: {current_row - 1}")
    print(f"Обработано файлов: {len(wop_files)}")
    
    return output_file

def copy_wop_to_excel_exactly():
    """
    Точное копирование: как пользователь вручную
    """
    print("=" * 60)
    print("ТОЧНОЕ КОПИРОВАНИЕ (КАК ВРУЧНУЮ)")
    print("=" * 60)
    
    # Получаем файлы
    wop_files = []
    for root, dirs, files in os.walk("."):
        for file in files:
            if file.lower().endswith('.wop'):
                wop_files.append(os.path.join(root, file))
    
    if not wop_files:
        print("Нет файлов .wop!")
        return None
    
    wop_files.sort()
    
    print(f"\nФайлов найдено: {len(wop_files)}")
    
    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Files"
    
    row = 1
    total_files = len(wop_files)
    
    for i, file_path in enumerate(wop_files, 1):
        filename = os.path.basename(file_path)
        print(f"[{i}/{total_files}] {filename}")
        
        # Читаем файл
        try:
            # Пробуем разные кодировки
            content = None
            for encoding in ['utf-8', 'windows-1251', 'cp1251', 'ascii']:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        content = f.read()
                    break
                except:
                    continue
            
            if content is None:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
            
            # Разделяем на строки
            lines = content.splitlines()
            
            # Записываем строки
            for line in lines:
                ws.cell(row=row, column=1, value=line)
                row += 1
            
            # Добавляем пустую строку между файлами
            if i < total_files:
                row += 1  # Одна пустая строка
                
        except Exception as e:
            print(f"  Ошибка: {e}")
            ws.cell(row=row, column=1, value=f"ERROR READING FILE: {filename}")
            row += 1
            if i < total_files:
                row += 1
    
    # Настраиваем
    ws.column_dimensions['A'].width = 200
    
    # Сохраняем
    output = "exact_copy.xlsx"
    wb.save(output)
    
    print(f"\n✅ Создан: {output}")
    print(f"📊 Строк: {row - 1}")
    
    return output

def show_preview():
    """Показать превью содержимого"""
    print("\n" + "=" * 60)
    print("ПРЕВЬЮ ФАЙЛОВ .wop:")
    print("=" * 60)
    
    wop_files = []
    for root, dirs, files in os.walk("."):
        for file in files:
            if file.lower().endswith('.wop'):
                wop_files.append(os.path.join(root, file))
    
    if not wop_files:
        print("Файлы не найдены")
        return
    
    wop_files.sort()
    
    for file_path in wop_files[:3]:  # Первые 3 файла
        filename = os.path.basename(file_path)
        print(f"\n{'='*40}")
        print(f"ФАЙЛ: {filename}")
        print(f"{'='*40}")
        
        try:
            content, _ = read_file_all_content(file_path)
            lines = content.split('\n')[:10]  # Первые 10 строк
            
            for i, line in enumerate(lines, 1):
                # Обрезаем длинные строки
                if len(line) > 100:
                    line = line[:100] + "..."
                print(f"{i:3}: {line}")
            
            if len(content.split('\n')) > 10:
                print("   ...")
            
            print(f"Всего строк: {len(content.split('\n'))}")
            
        except Exception as e:
            print(f"Ошибка чтения: {e}")

def create_test_files():
    """Создать тестовые файлы"""
    print("\nСоздаю тестовые файлы...")
    
    # Файл 1
    content1 = """W=92


Л1 ГОСТ 12.4.026-2015
Цвета сигнальные, знаки безопасности и разметка сигнальная 
(с изменением № 1, введенным в действие пост. Государственного
комитета по стандартизации Республики Беларусь от 07.07.2019 № 34)



 
N01   (Л1 Приложение Е, знак P01)
{+P01.WMF}
   Что означает это знак?
   1. Запрещается курить.
   2. Запрещается тушить окурки о строительные 
конструкции и элементы оборудования.
   3. Запрещается выбрасывать окурки.
   4. Место для тушения окурков.  

N02   (Л1 Приложение Е, знак P02)
{+P02.WMF}
   Что запрещает это знак?
   1. Пользоваться открытым огнем и курить.
   2. Пользоваться открытым огнем. (Курить можно,
если прикурить до входа на территорию, обозначенную этим знаком).
   3. Выбрасывать спички и окурки.
   4. Носить в карманах спички."""
    
    # Файл 2
    content2 = """Второй файл для теста

Строка 1 второго файла
Строка 2
Строка 3

Пустая строка выше"""
    
    # Создаем папку если нет
    os.makedirs("test_data", exist_ok=True)
    
    # Записываем файлы
    with open("test1.wop", "w", encoding="utf-8") as f:
        f.write(content1)
    
    with open(os.path.join("test_data", "test2.wop"), "w", encoding="utf-8") as f:
        f.write(content2)
    
    print("✅ Созданы тестовые файлы:")
    print("   - test1.wop")
    print("   - test_data/test2.wop")

def main():
    """Главная функция"""
    try:
        # Проверка библиотеки
        try:
            from openpyxl import Workbook
        except ImportError:
            print("Установите openpyxl: pip install openpyxl")
            return
        
        while True:
            print("\n" + "=" * 60)
            print("ГЛАВНОЕ МЕНЮ")
            print("=" * 60)
            print("1. Скопировать файлы в Excel (простой режим)")
            print("2. Скопировать файлы в Excel (точная копия)")
            print("3. Показать превью файлов")
            print("4. Создать тестовые файлы")
            print("5. Выход")
            print("-" * 60)
            
            choice = input("Выберите (1-5): ").strip()
            
            if choice == '1':
                copy_wop_to_excel_simple()
                input("\nEnter для продолжения...")
            
            elif choice == '2':
                copy_wop_to_excel_exactly()
                input("\nEnter для продолжения...")
            
            elif choice == '3':
                show_preview()
                input("\nEnter для продолжения...")
            
            elif choice == '4':
                create_test_files()
                input("\nEnter для продолжения...")
            
            elif choice == '5':
                print("\nВыход")
                break
            
            else:
                print("Неверный выбор")
    
    except KeyboardInterrupt:
        print("\nПрервано")
    except Exception as e:
        print(f"Ошибка: {e}")

# Самый простой вариант - одна функция
def ultra_simple():
    """
    УЛЬТРА-ПРОСТОЙ ВАРИАНТ
    Просто копирует все .wop файлы в Excel
    """
    import os
    from openpyxl import Workbook
    
    print("Собираю .wop файлы...")
    
    # Все .wop файлы в текущей папке и подпапках
    wop_files = []
    for root, dirs, files in os.walk("."):
        for file in files:
            if file.lower().endswith(".wop"):
                wop_files.append(os.path.join(root, file))
    
    if not wop_files:
        print("Файлы .wop не найдены!")
        return
    
    wop_files.sort()
    print(f"Найдено файлов: {len(wop_files)}")
    
    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    
    row = 1
    
    for file_path in wop_files:
        print(f"Обрабатываю: {os.path.basename(file_path)}")
        
        try:
            # Пробуем читать файл
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()
        except:
            try:
                with open(file_path, "r", encoding="windows-1251") as f:
                    content = f.read()
            except:
                try:
                    with open(file_path, "r", encoding="cp1251") as f:
                        content = f.read()
                except:
                    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                        content = f.read()
        
        # Разделяем на строки
        lines = content.split("\n")
        
        # Копируем в Excel
        for line in lines:
            ws.cell(row=row, column=1, value=line)
            row += 1
        
        # Пустая строка между файлами
        row += 1
    
    # Удаляем последнюю пустую строку
    if row > 1:
        row -= 1
    
    # Сохраняем
    ws.column_dimensions['A'].width = 150
    wb.save("ULTRA_SIMPLE_RESULT.xlsx")
    
    print(f"\n✅ ГОТОВО!")
    print(f"Файл: ULTRA_SIMPLE_RESULT.xlsx")
    print(f"Строк: {row - 1}")
    print(f"Файлов: {len(wop_files)}")

if __name__ == "__main__":
    # Просто запустите ultra_simple() для самого простого варианта
    # или main() для меню
    
    print("=" * 60)
    print("КОПИРОВАНИЕ .wop ФАЙЛОВ В EXCEL")
    print("=" * 60)
    print("\n1. Запустить простую версию (рекомендуется)")
    print("2. Запустить с меню")
    
    choice = input("\nВыберите (1 или 2): ").strip()
    
    if choice == '1':
        ultra_simple()
    else:
        main()
    
    input("\nНажмите Enter для выхода...")